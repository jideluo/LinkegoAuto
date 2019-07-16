# coding:utf-8
from bin.scriptUtils import utils
import xlsxwriter,xlrd
import openpyxl
import logging
# from ExcelContentProcess import ExcelClear,ExcleBorder,ExcelMean,ExcleTestResult,TestPersonnel


class WriteDate:
    def __init__(self):
        """
        初始化一个工作薄
        :return:
        """
        self.__workbook = xlsxwriter.Workbook(
            "{location}{time} .xlsx".format(location="./report/", time=utils.timestamp(),
                                                    ), {'strings_to_numbers': True})
        # self.__workbook = xlsxwriter.Workbook(
        #     "{location}{time}.xlsx".format(location="./chart/", time=utils.timestamp()), {'strings_to_numbers': True})

        self.format = self.__workbook.add_format({'num_format': 'yyyy-mm-dd %H.%M.%S'})
        self.format_head = self.__workbook.add_format({'bold': True})

    def __sheet(self, sheet_name):
        """
        创建工作表
        :return:
        """
        self.__worksheet = self.__workbook.add_worksheet(name=sheet_name)
        return self.__worksheet.set_column("A:Z", 20)

    def __write(self, heads, data):
        """
        写入数据头及数据,自适应数据数组行列
        :return:
        """
        for index, value in enumerate(heads):
            co = chr(66 + index)
            row = "{colum}{index}".format(colum=co, index=1)
            self.__worksheet.write("A1", "time", self.format_head)
            self.__worksheet.write(str(row), value, self.format_head)

        if len(data) == 1:
            for one, values in enumerate(data):
                for index, value in enumerate(values):
                    co = chr(65 + index)
                    row = "{colum}{index}".format(colum=co, index=2)
                    self.__worksheet.write(str(row), value)
        else:
            for more, values in enumerate(data):
                for index, value in enumerate(values):
                    co = chr(65 + index)
                    row = "{colum}{index}".format(colum=co, index=2 + more)
                    self.__worksheet.write(str(row), value)

    def __chart_series(self, sheetName, data):
        """
        制表数据
        :param sheetName:表名
        :return:
        """
        chartSeries = self.__workbook.add_chart({'type': 'line'})
        chartSeries.add_series({
            "name": '=%s!$B$1' % sheetName,  # 数据名
            "categories": '=' + sheetName + '!$A$2:$A$%s' % (len(data) + 1),  # 数据列长度
            "values": '=' + sheetName + '!$B$2:$B$%s' % (len(data) + 1),  # 数据列值
            "line": {'color': 'red'}
        })
        chartSeries.set_title({'name': sheetName})  # 图表标题
        chartSeries.set_x_axis({'name': u'时间'})  # x轴标题
        chartSeries.set_y_axis({'name': u'值'})  # y轴标题
        return chartSeries

    def __line_chart(self, sheetName, heads, data):
        """
        绘图
        :param sheetName:
        :return:
        """
        return self.__worksheet.insert_chart('%s2' % chr(len(heads) + 66), self.__chart_series(sheetName, data),
                                             {'x_offset': 10, 'y_offset': 10})

    def close(self):
        """
        关闭工作薄
        :return:
        """
        self.__workbook.close()

    def writeData(self, sheet_name, heads, data):
        """
        写入流程
        :param sheet_name:表名
        :param heads:数据头
        :param data:数据
        :return:
        """
        self.__sheet(sheet_name)
        self.__write(heads, data)
        self.__line_chart(sheet_name, heads, data)
        # self.close()  # 当前脚本调试时打开

class ReadDate:
    """
    读取文件AutoTestCase.xlsx

    """
    def __init__(self):
        self.xl = xlrd.open_workbook("report/AutoTestCase.xlsx");
        # self.xl = xlrd.open_workbook("../../report/AutoTestCase.xlsx");
    def read_app_name(self):
        """
        读取文件中APP包名
        :return:
        """
        self.table =self.xl.sheets()[1]
        logging.debug("APP包名：%s",self.table.row_values(0)[1])
        if(self.table.row_values(0)[1] is None):
            logging.debug("APP包名为空")
        else:
            return self.table.row_values(0)[1]
    def read_app_ip(self):
        """
        从文件读取手机IP地址
        :return:
        """
        self.table = self.xl.sheets()[1]
        logging.debug("IP地址：%s", self.table.row_values(4)[3])
        if (self.table.row_values(4)[3] is None):
            logging.debug("APP包名为空")
        else:
            return self.table.row_values(4)[3]
    def read_case(self):
        """
        从文件读取测试用例数据
        :return:
        """
        self.table =self.xl.sheets()[2]
        self.rows = self.table.nrows
        self.cols = self.table.ncols

        row_values = self.table.row_values(1)  # 获取第一行的内容，得到的是一个列表
        col_values = self.table.col_values(0)

        print(row_values)
        print(col_values)







        # def read(self):
    #     self.table =self.xl.sheets()[0]
    #     rows = self.table.nrows
    #     cols = self.table.ncols
    #     print(rows)
    #     print(cols)
    #     print(self.table)
        # Excel = openpyxl.load_workbook("PostTest.xlsx")  # 打开excel文件
        # # ExcelClear(Excel)
        # # print(Excel_File.get_sheet_names())#获取工作列表
        # # test = File.get_sheet_names()
        # # print(test.__len__())#获取工作列表长度
        # # print(test[0])#获取到第一个工作列表
        # sheet = Excel.get_sheet_by_name('post')
        # # print(sheet.title)
        # # sheet1 =File.get_active_sheet()
        # # print(sheet1.title)
        # maxColum = sheet.max_column
        # maxRow = sheet.max_row



if __name__ == '__main__':
    re =ReadDate()
    re.read_app_name()
    re.read_app_ip()
    re.read_case()
    # wd = WriteDate()
    # l = []
    # li = ['2016-01-15', '20008', '13656']
    # lo = ['2016-01-15', '20028', '13676', '3328']
    # lp = ['2016-01-15', '20028']
    # l.append(li)
    # l.append(lo)
    # l.append(lp)
    # heads = ["a", "b", "c"]
    # wd.writeData(sheet_name="test", heads=heads, data=l)
